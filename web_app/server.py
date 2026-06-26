"""Flask web server — SCCB Automation Web UI.

기존 sccb_app/jira_client.py, sccb_app/workflow.py를 그대로 재사용한다.
UI 레이어(tkinter)만 Flask + HTML/JS로 교체.
"""
import sys
import os
import io
import json
import queue
import threading
import concurrent.futures
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from flask import (
    Flask, request, jsonify, session, Response, send_file, render_template
)

# sccb_app 패키지를 import 경로에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from sccb_app.jira_client import JiraClient
from sccb_app.workflow import TransitionWorkflow

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "sccb-dev-secret-change-in-prod")

TZ = ZoneInfo("Asia/Seoul")

JQL_SCCB_NOT_TARGET = (
    '(category = AMHS_SW or category = AMHS) AND (status = "In Verification" or status = "Approval") '
    'AND ("개발 DR 완료일" >= 2022-01-01) AND ("SCCB 상태" = "미 대상") '
    'ORDER BY updated DESC'
)
JQL_SCCB_TARGET = (
    '(category = AMHS_SW or category = AMHS) AND (status = "In Verification" or status = "Approval") '
    'AND ("개발 DR 완료일" >= 2022-01-01) '
    'AND ("SCCB 상태" is EMPTY OR "SCCB 상태" = "SCCB 완료" OR "SCCB 상태" = "사전 SCCB 의뢰 완료") '
    'ORDER BY updated DESC'
)

# 진행 중인 검증 작업 (job_id → queue)
_validation_jobs: dict[str, queue.Queue] = {}
_jobs_lock = threading.Lock()


def _make_client() -> JiraClient:
    return JiraClient(
        base_url=session.get("base_url", ""),
        user=session.get("user", ""),
        password=session.get("password", ""),
        verify_ssl=session.get("verify_ssl", True),
        timeout=30,
    )


def _log_ts(msg: str) -> str:
    ts = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
    return f"[{ts}] {msg}"


# ─────────────────────────────────────────────
#  라우트
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) or {}
    session["base_url"] = (data.get("base_url") or "").rstrip("/")
    session["user"] = data.get("user", "")
    session["password"] = data.get("password", "")
    session["verify_ssl"] = bool(data.get("verify_ssl", True))
    try:
        jira = _make_client()
        me = jira.myself()
        name = (me.get("displayName") or "").strip()
        return jsonify({"ok": True, "name": name})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 401


@app.route("/api/search", methods=["POST"])
def api_search():
    data = request.get_json(force=True) or {}
    mode = data.get("mode", "not_target")
    jql = data.get("jql", "")
    max_results = int(data.get("max_results", 50))
    weekly_url = (data.get("weekly_url") or "").strip()

    try:
        jira = _make_client()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    logs = []
    weekly_keys: set[str] = set()

    try:
        if mode == "voc_complete":
            if not weekly_url:
                return jsonify({"ok": False, "error": "이번주 SCCB URL이 비어 있습니다."})
            weekly_keys = jira.get_weekly_sccb_issue_keys(weekly_url)
            logs.append(_log_ts(f"VOC 완료처리 대상 URL 파싱 완료 - {len(weekly_keys)}개 키"))
            if not weekly_keys:
                return jsonify({"ok": True, "issues": [], "total": 0, "logs": logs, "weekly_keys": []})
            key_list = sorted(weekly_keys)
            jql = f"key in ({','.join(key_list)}) ORDER BY key ASC"
            result_data = jira.search(jql, max(len(key_list), max_results))
        else:
            if weekly_url:
                try:
                    weekly_keys = jira.get_weekly_sccb_issue_keys(weekly_url)
                    logs.append(_log_ts(f"이번주 SCCB URL 파싱 완료 - {len(weekly_keys)}개 키"))
                except Exception as e:
                    logs.append(_log_ts(f"이번주 SCCB URL 파싱 실패 - {e}"))
            result_data = jira.search(jql, max_results)

        issues = result_data.get("issues", [])
        total = result_data.get("total", len(issues))

        rows = []
        for it in issues:
            f = it["fields"]
            key = it["key"]
            assignee_raw = (f.get("assignee") or {}).get("displayName", "")
            assignee = re.sub(r"\s*/.*$", "", assignee_raw)
            rows.append({
                "key": key,
                "summary": f.get("summary", ""),
                "status": (f.get("status") or {}).get("name", ""),
                "assignee": assignee,
                "duedate": f.get("duedate") or "",
                "in_weekly": key in weekly_keys,
                # 검증 컬럼: target 모드일 때 별도 스트림으로 채움
                "body_len": "...",
                "rollout": "...",
                "err_table": "...",
                "links": "...",
                "tcgen": "...",
                "aio_test": "...",
                "pr_merge": "...",
            })

        logs.append(_log_ts(f"Search 완료: {len(rows)}건 (전체 {total}건)"))
        return jsonify({
            "ok": True,
            "issues": rows,
            "total": total,
            "weekly_keys": list(weekly_keys),
            "logs": logs,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/validate/start", methods=["POST"])
def api_validate_start():
    """SCCB 대상 검증 작업을 시작하고 job_id를 반환한다."""
    data = request.get_json(force=True) or {}
    keys: list[str] = data.get("keys", [])
    if not keys:
        return jsonify({"ok": False, "error": "keys가 비어 있습니다."})

    # 세션 값을 캡처 (스레드에서 session 접근 불가)
    cfg = {
        "base_url": session.get("base_url", ""),
        "user": session.get("user", ""),
        "password": session.get("password", ""),
        "verify_ssl": session.get("verify_ssl", True),
    }

    job_id = f"{datetime.now(TZ).strftime('%Y%m%d%H%M%S%f')}"
    q: queue.Queue = queue.Queue()
    with _jobs_lock:
        _validation_jobs[job_id] = q

    def worker():
        try:
            jira = JiraClient(
                base_url=cfg["base_url"],
                user=cfg["user"],
                password=cfg["password"],
                verify_ssl=cfg["verify_ssl"],
                timeout=30,
            )

            # 배치 core 조회
            try:
                core_cache = jira.get_issues_core_batch(keys)
            except Exception:
                core_cache = {}

            def fetch_one(k: str):
                try:
                    core = core_cache.get(k) or jira.get_issue_core(k)
                    fields = (core.get("fields") or {})
                    desc = fields.get("description") or ""
                    issuelinks = fields.get("issuelinks") or []
                    issue_id = core.get("id") or ""

                    result = {"key": k}

                    # 빠른 항목
                    try:
                        result["rollout"] = jira.get_design_rollout_status(k, desc=desc)
                    except Exception:
                        result["rollout"] = "ERR"
                    try:
                        result["err_table"] = jira.get_error_table_status(k, desc=desc)
                    except Exception:
                        result["err_table"] = "ERR"
                    try:
                        missing = (jira.get_link_validation(k, links=issuelinks).get("missing") or [])
                        result["links"] = "OK" if not missing else "Missing: " + ",".join(missing)
                    except Exception:
                        result["links"] = "ERR"

                    # 느린 항목
                    try:
                        bl = jira.get_body_length_string_from_ui(k)
                        result["body_len"] = bl if bl else "확인불가"
                    except Exception:
                        result["body_len"] = "ERR"
                    try:
                        tc = jira.get_tc_generation_check(k)
                        result["tcgen"] = "OK" if tc.get("ok") else "FAIL"
                    except Exception:
                        result["tcgen"] = "ERR"
                    try:
                        aio = jira.get_aio_test_validation(k)
                        result["aio_test"] = aio.get("status") or "ERR"
                    except Exception:
                        result["aio_test"] = "ERR"
                    try:
                        pr = jira.get_pr_merge_status(k, issue_id)
                        result["pr_merge"] = pr if pr is not None else "ERR"
                    except Exception:
                        result["pr_merge"] = "ERR"

                    q.put({"type": "update", "data": result})
                except Exception as e:
                    q.put({"type": "update", "data": {
                        "key": k,
                        "body_len": "ERR", "rollout": "ERR", "err_table": "ERR",
                        "links": "ERR", "tcgen": "ERR", "aio_test": "ERR", "pr_merge": "ERR",
                    }})

            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
                futures = [ex.submit(fetch_one, k) for k in keys]
                concurrent.futures.wait(futures)
        finally:
            q.put({"type": "done"})

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"ok": True, "job_id": job_id})


@app.route("/api/validate/stream/<job_id>")
def api_validate_stream(job_id: str):
    """Server-Sent Events로 검증 결과를 스트리밍한다."""
    with _jobs_lock:
        q = _validation_jobs.get(job_id)
    if q is None:
        return Response("data: {\"type\":\"error\",\"msg\":\"job not found\"}\n\n",
                        mimetype="text/event-stream")

    def generate():
        while True:
            try:
                msg = q.get(timeout=60)
            except queue.Empty:
                yield "data: {\"type\":\"heartbeat\"}\n\n"
                continue
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg.get("type") == "done":
                with _jobs_lock:
                    _validation_jobs.pop(job_id, None)
                break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/transition/approval", methods=["POST"])
def api_transition_approval():
    data = request.get_json(force=True) or {}
    keys: list[str] = data.get("keys", [])
    if not keys:
        return jsonify({"ok": False, "error": "keys가 비어 있습니다."})

    logs = []
    try:
        jira = _make_client()
        wf = TransitionWorkflow(
            jira_client=jira,
            log_fn=lambda msg: logs.append(_log_ts(msg)),
            sccb_mode_getter=lambda: data.get("mode", "not_target"),
        )
        logs.append(_log_ts(f"Approval 전이 시작: {keys}"))
        statuses = {}
        for k in keys:
            try:
                wf.process_issue_to_approval(k)
                try:
                    statuses[k] = jira.get_issue_status(k)
                except Exception:
                    pass
            except Exception as e:
                logs.append(_log_ts(f"{k}: 실패 - {e}"))
        logs.append(_log_ts("Approval 전이 종료"))
        return jsonify({"ok": True, "logs": logs, "statuses": statuses})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "logs": logs}), 500


@app.route("/api/transition/complete", methods=["POST"])
def api_transition_complete():
    data = request.get_json(force=True) or {}
    keys: list[str] = data.get("keys", [])
    mode = data.get("mode", "not_target")
    if not keys:
        return jsonify({"ok": False, "error": "keys가 비어 있습니다."})

    logs = []
    try:
        jira = _make_client()
        wf = TransitionWorkflow(
            jira_client=jira,
            log_fn=lambda msg: logs.append(_log_ts(msg)),
            sccb_mode_getter=lambda: mode,
        )
        logs.append(_log_ts(f"처리 시작: {keys}"))
        statuses = {}
        for k in keys:
            try:
                if mode == "voc_complete":
                    wf.process_voc_linked_issues_from_parent(k)
                else:
                    wf.process_issue_to_complete(k)
                try:
                    statuses[k] = jira.get_issue_status(k)
                except Exception:
                    pass
            except Exception as e:
                logs.append(_log_ts(f"{k}: 실패 - {e}"))
        logs.append(_log_ts("처리 종료"))
        return jsonify({"ok": True, "logs": logs, "statuses": statuses})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "logs": logs}), 500


@app.route("/api/create_sccb_page", methods=["POST"])
def api_create_sccb_page():
    data = request.get_json(force=True) or {}
    weekly_url = (data.get("weekly_url") or "").strip()
    if not weekly_url:
        return jsonify({"ok": False, "error": "이번주 SCCB URL을 입력하세요."})
    try:
        jira = _make_client()
        result = jira.create_next_week_confluence_page_from_url(weekly_url)
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/create_meeting_page", methods=["POST"])
def api_create_meeting_page():
    data = request.get_json(force=True) or {}
    source_url = (data.get("source_url") or "").strip()
    if not source_url:
        return jsonify({"ok": False, "error": "회의록 원본 URL을 입력하세요."})
    try:
        jira = _make_client()
        result = jira.create_next_week_meeting_minutes_page_from_url(source_url)
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/export_excel", methods=["POST"])
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True) or {}
    rows_data: list[dict] = data.get("rows", [])
    if not rows_data:
        return jsonify({"ok": False, "error": "내보낼 데이터가 없습니다."}), 400

    def calc_result(r: dict) -> str:
        rollout = (r.get("rollout") or "")
        err_table = (r.get("err_table") or "")
        links = (r.get("links") or "")
        tcgen = (r.get("tcgen") or "")
        aio_test = (r.get("aio_test") or "")
        pr_merge = (r.get("pr_merge") or "")
        if "OK" not in rollout.upper():
            return "공란"
        if "OK" not in err_table.upper():
            return "FAIL"
        if "OK" not in links.upper():
            return "FAIL"
        if "OK" not in tcgen.upper():
            return "FAIL"
        if "OK" not in aio_test.upper():
            return "FAIL"
        if "MERGED" not in pr_merge.upper():
            if "N/A" in pr_merge.upper():
                return "N/A"
            return "FAIL"
        return "OK"

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    header = ["RESULT", "KEY", "SUMMARY", "본문 길이", "횡전개 표", "연관 에러",
              "이슈연결", "LLM TC 생성", "AIO Test", "P/R 병합", "STATUS", "ASSIGNEE", "기한일"]
    ws.append(header)
    header_font = Font(bold=True)
    for ci in range(1, len(header) + 1):
        ws.cell(row=1, column=ci).font = header_font

    for r in rows_data:
        result = calc_result(r)
        ws.append([
            result,
            r.get("key", ""),
            r.get("summary", ""),
            r.get("body_len", ""),
            r.get("rollout", ""),
            r.get("err_table", ""),
            r.get("links", ""),
            r.get("tcgen", ""),
            r.get("aio_test", ""),
            r.get("pr_merge", ""),
            r.get("status", ""),
            r.get("assignee", ""),
            r.get("duedate", ""),
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows_data)+1}"
    ws.freeze_panes = "A2"
    widths = [10, 16, 70, 12, 12, 12, 28, 12, 12, 10, 18, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
    filename = f"sccb_validation_{ts}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
