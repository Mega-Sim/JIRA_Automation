import re
import tkinter as tk
from tkinter import filedialog
from tkinter import font as tkfont
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from tkinter import ttk

import threading
import webbrowser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from zoneinfo import ZoneInfo

DEFAULT_BASE_URL = "https://jira-stms.semes.com:18080"
TZ = ZoneInfo("Asia/Seoul")
JQL_SCCB_NOT_TARGET = (
    '(category = AMHS_SW or category = AMHS) AND (status = "In Verification" or status = "Approval") '
    'AND ("개발 DR 완료일" >= 2022-01-01) AND ("SCCB 상태" = "미 대상") '
    'ORDER BY updated DESC'
)
JQL_SCCB_TARGET = (
    '(category = AMHS_SW or category = AMHS) AND (status = "In Verification" or status = "Approval") '
    'AND ("개발 DR 완료일" >= 2022-01-01) '
    'AND ("SCCB 상태" is EMPTY OR "SCCB 상태" = "SCCB 완료" OR "SCCB 상태" = "사전 SCCB 의뢰 완료") '
    'ORDER BY updated DESC'
)
# (본문 길이는 Jira에서 제공하는 필드 값을 그대로 사용)
from .jira_client import JiraClient
from .workflow import TransitionWorkflow


class JiraSccbApp(tb.Window):
    def __init__(self):
        super().__init__(themename="darkly")
        # 결과 그리드 가독성 개선: 폰트/행높이 10% 확대 (v22)
        try:
            style = getattr(self, 'style', None) or tb.Style()
            base_font = tkfont.nametofont('TkDefaultFont')
            base_size = int(base_font.cget('size'))
            new_size = max(base_size + 1, int(round(base_size * 1.1)))
            family = base_font.actual().get('family', 'TkDefaultFont')
            style.configure('Treeview', font=(family, new_size))
            style.configure('Treeview.Heading', font=(family, new_size))
            try:
                cur_rh = style.lookup('Treeview', 'rowheight')
                cur_rh = int(cur_rh) if str(cur_rh).isdigit() else 20
            except Exception:
                cur_rh = 20
            style.configure('Treeview', rowheight=int(round(cur_rh * 1.1)))
        except Exception:
            pass
        self.title("Jira SCCB Automation v96")
        self.geometry("1500x900")

        self.base_url_var = tk.StringVar(value=DEFAULT_BASE_URL)
        self.user_var = tk.StringVar()
        self.pass_var = tk.StringVar()
        self.verify_ssl_var = tk.BooleanVar(value=True)
        self.jql_var = tk.StringVar(value="")
        self.weekly_sccb_url_var = tk.StringVar(value="")
        self.max_results_var = tk.IntVar(value=50)

        self.selected = {}
        self.issue_status = {}
        self.iid_by_key = {}
        self.select_all_var = tk.IntVar(value=0)
        self.sccb_mode = "not_target"  # 'not_target' or 'target' or 'voc_complete'
        self.weekly_sccb_keys = set()
        self.body_len_field_id = None
        self._sort_states = {}
        self._column_titles = {}

        self._build_ui()

    def _make_client(self) -> JiraClient:
        return JiraClient(
            base_url=self.base_url_var.get(),
            user=self.user_var.get(),
            password=self.pass_var.get(),
            verify_ssl=self.verify_ssl_var.get(),
            timeout=30
        )

    def _build_ui(self):
        frm = tb.Frame(self, padding=12)
        frm.pack(fill=BOTH, expand=YES)

        BTN_W = 16
        LABEL_W = 10

        conn = tb.Labelframe(frm, text="접속 정보", padding=10)
        conn.pack(fill=X)

        # 2단(좌/우) 레이아웃: 좌측=BaseURL/ID/PW/로그인, 우측=SSL/MaxResults
        left = tb.Frame(conn)
        left.grid(row=0, column=0, rowspan=2, sticky=EW, padx=(0, 16))

        right = tb.Frame(conn)
        right.grid(row=0, column=1, rowspan=2, sticky=E)

        conn.columnconfigure(0, weight=1)
        conn.columnconfigure(1, weight=0)

        # --- Left (Base URL + Credentials) ---
        # Base URL은 가로로 늘어나도, ID/PW 옆 로그인 영역이 밀리지 않도록 블록을 분리한다.
        left.columnconfigure(0, weight=1)

        base_row = tb.Frame(left)
        base_row.grid(row=0, column=0, sticky=EW)
        base_row.columnconfigure(1, weight=1)

        tb.Label(base_row, text="Base URL", width=LABEL_W, anchor="e").grid(row=0, column=0, sticky=E)
        tb.Entry(base_row, textvariable=self.base_url_var).grid(row=0, column=1, sticky=EW, padx=6)

        cred = tb.Frame(left)
        cred.grid(row=1, column=0, sticky=W, pady=(8, 0))

        tb.Label(cred, text="ID", width=LABEL_W, anchor="e").grid(row=0, column=0, sticky=E)
        self.id_entry = tb.Entry(cred, textvariable=self.user_var, width=30)
        self.id_entry.grid(row=0, column=1, sticky=W, padx=6)

        # 로그인 상태/버튼: ID/PW 입력칸 '바로 옆' 고정
        self.login_lbl = tb.Label(cred, text="로그인: 미확인", anchor=W)
        self.login_lbl.grid(row=0, column=2, sticky=W, padx=(10, 0))

        tb.Label(cred, text="PW", width=LABEL_W, anchor="e").grid(row=1, column=0, sticky=E, pady=(6, 0))
        self.pw_entry = tb.Entry(cred, textvariable=self.pass_var, width=30, show="*")
        self.pw_entry.grid(row=1, column=1, sticky=W, padx=6, pady=(6, 0))

        tb.Button(cred, text="Login 확인", bootstyle=SUCCESS, command=self.on_check_login, width=BTN_W).grid(
            row=1, column=2, sticky=W, padx=(10, 0), pady=(6, 0)
        )

        # Enter로 로그인 (ID/PW 입력칸에서)
        self.id_entry.bind("<Return>", lambda e: self.on_check_login())
        self.pw_entry.bind("<Return>", lambda e: self.on_check_login())
        self.id_entry.focus_set()

        # --- Right (Options) ---
        tb.Checkbutton(right, text="SSL verify", variable=self.verify_ssl_var).grid(row=0, column=0, sticky=W)
        opt_row = tb.Frame(right)
        opt_row.grid(row=1, column=0, sticky=W, pady=(6, 0))
        tb.Label(opt_row, text="Max Results").pack(side=LEFT)
        tb.Spinbox(opt_row, from_=1, to=500, textvariable=self.max_results_var, width=8).pack(side=LEFT, padx=(8, 0))

        # conn 자체는 좌측 확장, 우측은 고정
        conn.columnconfigure(0, weight=1)
        conn.columnconfigure(1, weight=0)

        preset = tb.Frame(frm, padding=10)
        preset.pack(fill=X, pady=(12, 0))
        tb.Button(preset, text="SCCB 미대상", bootstyle=WARNING, command=self.on_sccb_not_target, width=BTN_W).pack(side=LEFT)
        tb.Button(preset, text="SCCB 대상", bootstyle=PRIMARY, command=self.on_sccb_target, width=BTN_W).pack(side=LEFT, padx=8)
        tb.Button(preset, text="VOC 완료처리", bootstyle=INFO, command=self.on_voc_complete, width=BTN_W).pack(side=LEFT)
        tb.Button(preset, text="SCCB Page 생성", bootstyle=SUCCESS, command=self.on_create_sccb_page, width=BTN_W).pack(side=LEFT, padx=(8, 0))
        tb.Button(preset, text="회의록 Page 생성", bootstyle=SUCCESS, command=self.on_create_meeting_page, width=BTN_W).pack(side=LEFT, padx=(8, 0))

        weekly_url = tb.Labelframe(
            frm,
            text="이번주 SCCB URL (이전 주 Page 포맷 복제 후 다음 주 URL로 자동 변경)",
            padding=10,
        )
        weekly_url.pack(fill=X, pady=(0, 0))
        tb.Entry(weekly_url, textvariable=self.weekly_sccb_url_var).pack(fill=X)

        jql = tb.Labelframe(frm, text="JQL", padding=10)
        jql.pack(fill=X, pady=(10, 0))
        tb.Entry(jql, textvariable=self.jql_var).pack(fill=X)

        btn_row = tb.Frame(frm, padding=10)
        btn_row.pack(fill=X, pady=10)

        tb.Button(btn_row, text="Search", bootstyle=SECONDARY, command=self.on_search, width=BTN_W).pack(side=LEFT)
        tb.Button(btn_row, text="Open Selected", bootstyle=INFO, command=self.open_selected_issue, width=BTN_W).pack(side=LEFT, padx=8)
        tb.Button(btn_row, text="선택 이슈 Approval", bootstyle=WARNING, command=self.on_process_selected_approval, width=BTN_W).pack(side=LEFT, padx=8)
        tb.Button(btn_row, text="선택 이슈 Complete", bootstyle=DANGER, command=self.on_process_selected, width=BTN_W).pack(side=LEFT, padx=8)
        tb.Button(btn_row, text="Export Excel", bootstyle=PRIMARY, command=self.on_export_excel, width=BTN_W).pack(side=LEFT, padx=8)

        self.status_lbl = tb.Label(btn_row, text="")
        self.status_lbl.pack(side=RIGHT)

        res = tb.Labelframe(frm, text="결과", padding=10)
        res.pack(fill=BOTH, expand=YES)

        # 전체 선택 체크박스 헤더 (첫 번째 컬럼 정렬)
        hdr = tb.Frame(res)
        hdr.pack(fill=X)
        self.select_all_chk = tb.Checkbutton(
            hdr,
            variable=self.select_all_var,
            bootstyle=SUCCESS,
            command=self._on_toggle_select_all
        )
        # Tree 첫 컬럼(sel) 폭에 맞춰 배치
        self.select_all_chk.pack(side=LEFT, padx=(8, 0))

        cols = ("sel", "key", "summary", "body_len", "rollout", "err_table", "links", "tcgen", "aio_test", "pr_merge", "status", "assignee", "duedate")
        self.tree = tb.Treeview(res, columns=cols, show="headings", height=16)
        for c, t in [
            ("sel", "선택"),
            ("key", "KEY"),
            ("summary", "SUMMARY"),
            ("body_len", "본문 길이"),
            ("rollout", "설계횡전개"),
            ("err_table", "연관 에러"),
            ("links", "이슈연결 상태"),
            ("tcgen", "LLM TC 생성"),
            ("aio_test", "AIO Test"),
            ("pr_merge", "P/R 병합"),
            ("status", "STATUS"),
            ("assignee", "ASSIGNEE"),
            ("duedate", "기한일"),
        ]:
            self._column_titles[c] = t
            self.tree.heading(c, text=t, command=lambda col=c: self._sort_tree_by_column(col))

        # 초기 화면에서는 모든 열이 보이도록 기본 폭을 줄이되,
        # 사용자가 드래그로 조절한 폭이 자동으로 되돌아가지 않도록 전 컬럼 stretch=False 유지.
        # SUMMARY는 길어도 잘려도 되므로 기본 폭을 가장 작게 잡고,
        # 값이 짧은 ASSIGNEE/STATUS 등은 실제 표시 길이에 맞춰 축소한다.
        self.tree.column("sel", width=42, minwidth=36, anchor=CENTER, stretch=False)
        self.tree.column("key", width=135, minwidth=100, stretch=False)
        self.tree.column("summary", width=170, minwidth=90, stretch=False)
        self.tree.column("body_len", width=76, minwidth=70, anchor=CENTER, stretch=False)
        self.tree.column("rollout", width=84, minwidth=78, anchor=CENTER, stretch=False)
        self.tree.column("err_table", width=84, minwidth=78, anchor=CENTER, stretch=False)
        self.tree.column("links", width=110, minwidth=95, anchor=CENTER, stretch=False)
        self.tree.column("tcgen", width=100, minwidth=95, anchor=CENTER, stretch=False)
        self.tree.column("aio_test", width=90, minwidth=82, anchor=CENTER, stretch=False)
        self.tree.column("pr_merge", width=82, minwidth=76, anchor=CENTER, stretch=False)
        self.tree.column("status", width=95, minwidth=88, stretch=False)
        self.tree.column("assignee", width=82, minwidth=72, stretch=False)
        self.tree.column("duedate", width=95, minwidth=92, anchor=CENTER, stretch=False)

        vsb = tb.Scrollbar(res, orient=VERTICAL, command=self.tree.yview)
        hsb = tb.Scrollbar(res, orient=HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        
        # 스크롤바를 먼저 배치해야 올바른 위치에 감
        vsb.pack(side=RIGHT, fill=Y)
        hsb.pack(side=BOTTOM, fill=X)
        self.tree.pack(side=LEFT, fill=BOTH, expand=YES)

        # zebra stripe
        self.tree.tag_configure("odd",  background="#2b2b2b")
        self.tree.tag_configure("even", background="#1e1e1e")
        self.tree.tag_configure("odd_match", background="#2b2b2b", foreground="#5cb85c")
        self.tree.tag_configure("even_match", background="#1e1e1e", foreground="#5cb85c")


        # Treeview style로 세로/가로 구분선 표시
        # ttkbootstrap darkly 테마 위에 덮어쓰기
        try:
            s = tb.Style()
            s.configure("Treeview",
                        rowheight=s.lookup("Treeview", "rowheight") or 22)
            # 헤더/셀 경계선: Treeview 자체 option
            self.tree.configure(style="Treeview")
        except Exception:
            pass

        def _apply_zebra():
            """Zebra stripe 적용 + 주간 SCCB 키 매칭 행 강조"""
            weekly_keys = getattr(self, "weekly_sccb_keys", set()) or set()
            for i, iid in enumerate(self.tree.get_children("")):
                stripe = "even" if i % 2 == 0 else "odd"
                values = self.tree.item(iid, "values")
                key = (values[1] if values and len(values) > 1 else "")
                if key and key in weekly_keys:
                    self.tree.item(iid, tags=[f"{stripe}_match"])
                else:
                    self.tree.item(iid, tags=[stripe])
        self._apply_zebra = _apply_zebra
        self._redraw_grid = lambda *a: None  # 하위 호환용 빈 함수

        self.tree.bind("<Button-1>", self._on_tree_click, add=True)

        logf = tb.Labelframe(frm, text="처리 로그", padding=10)
        logf.pack(fill=BOTH, pady=(10, 0))
        self.log_text = tk.Text(logf, height=10)
        self.log_text.pack(fill=BOTH, expand=YES)

        # Text 태그(부분 색상) 설정
        self.log_text.tag_configure("tag_fail", foreground="red")
        self.log_text.tag_configure("tag_len_low", foreground="red")

    def _update_heading_sort_indicators(self, active_col=None, descending=False):
        for col, title in self._column_titles.items():
            suffix = ""
            if col == active_col:
                suffix = " ▼" if descending else " ▲"
            self.tree.heading(col, text=title + suffix, command=lambda c=col: self._sort_tree_by_column(c))

    def _parse_sort_value(self, col: str, value):
        s = "" if value is None else str(value).strip()
        if col == "sel":
            if s == "☑":
                return (0, "")
            if s == "☐":
                return (1, "")
            return (2, s)

        if col == "body_len":
            m = re.search(r"(\d+)", s)
            return (0, int(m.group(1))) if m else (1, s.lower())

        if col == "duedate":
            for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return (0, datetime.strptime(s, fmt))
                except Exception:
                    pass
            return (1, s.lower())

        if col == "aio_test":
            m = re.match(r"^(OK|FAIL)\((\d+)/(\d+|-)\)$", s, re.IGNORECASE)
            if m:
                status_rank = 0 if m.group(1).upper() == "OK" else 1
                required = -1 if m.group(3) == "-" else int(m.group(3))
                return (0, status_rank, int(m.group(2)), required, s.lower())
            if s.upper().startswith("ERR"):
                return (1, 0, 0, 0, s.lower())
            return (2, s.lower())

        if col in ("rollout", "err_table", "tcgen"):
            up = s.upper()
            if up == "OK":
                return (0, s.lower())
            if up == "FAIL":
                return (1, s.lower())
            if up == "ERR":
                return (2, s.lower())
            if s == "...":
                return (3, s.lower())
            return (4, s.lower())

        if col == "pr_merge":
            up = s.upper()
            if up == "MERGED":
                return (0, s.lower())
            if up.startswith("OPEN"):
                return (1, s.lower())
            if up == "NONE":
                return (2, s.lower())
            if up.startswith("ERR"):
                return (3, s.lower())
            if s == "...":
                return (4, s.lower())
            return (5, s.lower())

        return s.lower()

    def _sort_tree_by_column(self, col: str):
        children = list(self.tree.get_children(""))
        if not children:
            return

        descending = not self._sort_states.get(col, False)
        self._sort_states = {col: descending}

        items = []
        for iid in children:
            value = self.tree.set(iid, col)
            items.append((self._parse_sort_value(col, value), iid))

        items.sort(key=lambda x: x[0], reverse=descending)

        for idx, (_, iid) in enumerate(items):
            self.tree.move(iid, "", idx)

        self._update_heading_sort_indicators(active_col=col, descending=descending)
        self._apply_zebra()

    def log(self, msg):
        ts = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

        # 이번에 삽입되는 라인의 시작 index를 확보한 뒤,
        # 삽입 후 해당 구간에서 "FAIL" / "N 자" (N<300) 만 부분 색상 처리한다.
        start_idx = self.log_text.index(END)

        line = f"[{ts}] {msg}\n"
        self.log_text.insert(END, line)

        end_idx = self.log_text.index(END)

        # 1) "FAIL" 부분만 빨간색
        search_from = start_idx
        while True:
            pos = self.log_text.search("FAIL", search_from, stopindex=end_idx, nocase=False)
            if not pos:
                break
            pos_end = f"{pos}+{len('FAIL')}c"
            self.log_text.tag_add("tag_fail", pos, pos_end)
            search_from = pos_end

        # 2) 본문 길이: "<숫자> 자" 형태를 찾아 숫자<300이면 해당 구간만 빨간색
        #    예) "299 자", "  12자" 등
        seg = self.log_text.get(start_idx, end_idx)
        for m in re.finditer(r"(\d+)\s*자", seg):
            n = int(m.group(1))
            if n < 300:
                a = m.start()
                b = m.end()
                a_idx = f"{start_idx}+{a}c"
                b_idx = f"{start_idx}+{b}c"
                self.log_text.tag_add("tag_len_low", a_idx, b_idx)

        self.log_text.see(END)

    def set_status(self, msg):
        self.status_lbl.config(text=msg)

    def _on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        if col != "#1":
            return
        row = self.tree.identify_row(event.y)
        if not row:
            return
        vals = self.tree.item(row, "values")
        key = vals[1]
        self.selected[key] = not self.selected.get(key, False)
        self.tree.set(row, "sel", "☑" if self.selected[key] else "☐")
        self._refresh_select_all_state()

    def _get_first_selected_key(self):
        for k, v in self.selected.items():
            if v:
                return k
        return None

    def _on_toggle_select_all(self):
        want = bool(self.select_all_var.get())
        for key, iid in self.iid_by_key.items():
            self.selected[key] = want
            self.tree.set(iid, "sel", "☑" if want else "☐")

    def _refresh_select_all_state(self):
        if not self.selected:
            self.select_all_var.set(0)
            return
        vals = list(self.selected.values())
        if all(vals):
            self.select_all_var.set(1)
        elif not any(vals):
            self.select_all_var.set(0)
        else:
            # partial: leave unchecked
            self.select_all_var.set(0)

    def on_sccb_not_target(self):
        self.sccb_mode = "not_target"
        self.jql_var.set(JQL_SCCB_NOT_TARGET)
        self.on_search()

    def on_sccb_target(self):
        self.sccb_mode = "target"
        self.jql_var.set(JQL_SCCB_TARGET)
        self.on_search()

    def on_voc_complete(self):
        weekly_url = (self.weekly_sccb_url_var.get() or "").strip()
        if not weekly_url:
            Messagebox.show_warning("이번주 SCCB URL을 입력하세요.", title="경고")
            return
        self.sccb_mode = "voc_complete"
        self.on_search()

    def _create_next_week_page_from_url(
        self,
        label: str,
        page_url: str,
        create_method_name: str,
        update_url_var=None,
    ):
        source_url = (page_url or "").strip()
        if not source_url:
            Messagebox.show_warning(f"{label} 원본 URL을 입력하세요.", title="경고")
            return

        def worker():
            try:
                self.log(f"{label} 이전 주 페이지 포맷 복제 시작 - 원본: {source_url}")
                jira = self._make_client()
                create_method = getattr(jira, create_method_name)
                result = create_method(source_url)
                source_title = result.get("source_title") or ""
                title = result.get("title") or ""
                url = result.get("url") or ""
                created = bool(result.get("created"))

                if created:
                    msg = f"{label} 생성 완료: {title}"
                else:
                    msg = f"{label} 이미 존재: {title}"
                if source_title:
                    msg += f" (원본: {source_title})"
                yearly_parent_title = result.get("yearly_parent_title") or ""
                if yearly_parent_title:
                    if result.get("yearly_parent_created"):
                        msg += f" / 상위 페이지 생성: {yearly_parent_title}"
                    else:
                        msg += f" / 상위 페이지 사용: {yearly_parent_title}"
                if url:
                    msg += f" / {url}"
                self.log(msg)

                def show_done():
                    if update_url_var is not None and url:
                        update_url_var.set(url)
                    Messagebox.show_info(msg, title="완료")
                    if url:
                        webbrowser.open(url)

                self.after(0, show_done)
            except Exception as e:
                self.log(f"{label} 생성 실패 - {e}")
                self.after(
                    0,
                    lambda: Messagebox.show_error(
                        f"{label} 생성 실패\n{e}",
                        title="오류",
                    ),
                )

        threading.Thread(target=worker, daemon=True).start()

    def on_create_sccb_page(self):
        self._create_next_week_page_from_url(
            "SCCB Page",
            self.weekly_sccb_url_var.get(),
            "create_next_week_confluence_page_from_url",
            update_url_var=self.weekly_sccb_url_var,
        )

    def on_create_meeting_page(self):
        """회의록 원본 URL을 팝업에서 받은 뒤 다음 주 회의록을 생성한다.

        메인 화면은 이번주 SCCB URL만 유지한다. 회의록은 매번 다른 원본을
        지정할 수 있도록 생성 버튼을 눌렀을 때만 URL 입력 팝업을 띄운다.
        """
        dialog = tb.Toplevel(self)
        dialog.title("회의록 원본 URL 입력")
        dialog.transient(self)
        dialog.resizable(False, False)
        dialog.grab_set()

        container = tb.Frame(dialog, padding=16)
        container.pack(fill=BOTH, expand=YES)

        tb.Label(
            container,
            text="이번주 회의록 페이지 URL을 입력하세요.",
            anchor=W,
        ).pack(fill=X)

        url_var = tk.StringVar(value="")
        entry = tb.Entry(container, textvariable=url_var, width=82)
        entry.pack(fill=X, pady=(8, 12))

        button_row = tb.Frame(container)
        button_row.pack(fill=X)

        def close_dialog():
            try:
                dialog.grab_release()
            except tk.TclError:
                pass
            dialog.destroy()

        def submit(event=None):
            source_url = (url_var.get() or "").strip()
            if not source_url:
                Messagebox.show_warning(
                    "회의록 원본 URL을 입력하세요.",
                    title="경고",
                )
                entry.focus_set()
                return "break" if event is not None else None

            close_dialog()
            self._create_next_week_page_from_url(
                "회의록 Page",
                source_url,
                "create_next_week_meeting_minutes_page_from_url",
            )
            return "break" if event is not None else None

        tb.Button(
            button_row,
            text="확인",
            bootstyle=SUCCESS,
            command=submit,
            width=12,
        ).pack(side=RIGHT)
        tb.Button(
            button_row,
            text="취소",
            bootstyle=SECONDARY,
            command=close_dialog,
            width=12,
        ).pack(side=RIGHT, padx=(0, 8))

        entry.bind("<Return>", submit)
        dialog.bind("<Escape>", lambda event: close_dialog())
        dialog.protocol("WM_DELETE_WINDOW", close_dialog)
        entry.focus_set()

    def on_check_login(self):
        def worker():
            try:
                jira = self._make_client()
                me = jira.myself()
                name = (me.get("displayName") or "").strip()
                msg = f"로그인 성공: {name}" if name else "로그인 성공"
                self.login_lbl.config(text=msg)
                self.log(msg)
            except Exception as e:
                self.login_lbl.config(text="로그인 실패")
                self.log(f"로그인 실패 - {e}")
        threading.Thread(target=worker, daemon=True).start()

    def on_search(self):
        if self.sccb_mode != "voc_complete" and not self.jql_var.get():
            Messagebox.show_warning("JQL이 비어 있습니다.", title="경고")
            return

        self.tree.delete(*self.tree.get_children())
        self.selected = {}
        self.issue_status = {}
        self.iid_by_key = {}
        self.select_all_var.set(0)


        def worker():
            try:
                jira = self._make_client()
                weekly_url = (self.weekly_sccb_url_var.get() or "").strip()
                weekly_keys = set()
                data = None

                if self.sccb_mode == "voc_complete":
                    if not weekly_url:
                        self.weekly_sccb_keys = set()
                        self.log("VOC 완료처리 실패 - 이번주 SCCB URL이 비어 있습니다.")
                        self.set_status("0 issues")
                        return
                    try:
                        weekly_keys = jira.get_weekly_sccb_issue_keys(weekly_url)
                        self.weekly_sccb_keys = weekly_keys
                        self.log(f"VOC 완료처리 대상 URL 파싱 완료 - {len(weekly_keys)}개 키")
                    except Exception as e:
                        self.weekly_sccb_keys = set()
                        self.log(f"VOC 완료처리 대상 URL 파싱 실패 - {e}")
                        self.set_status("0 issues")
                        return

                    if not weekly_keys:
                        self.log("VOC 완료처리 대상 이슈가 없습니다.")
                        self.set_status("0 issues")
                        return

                    key_list = sorted(weekly_keys)
                    jql = f"key in ({','.join(key_list)}) ORDER BY key ASC"
                    data = jira.search(jql, max(len(key_list), self.max_results_var.get()))
                else:
                    if weekly_url:
                        try:
                            weekly_keys = jira.get_weekly_sccb_issue_keys(weekly_url)
                            self.weekly_sccb_keys = weekly_keys
                            self.log(f"이번주 SCCB URL 파싱 완료 - {len(weekly_keys)}개 키")
                        except Exception as e:
                            self.weekly_sccb_keys = set()
                            self.log(f"이번주 SCCB URL 파싱 실패 - {e}")
                    else:
                        self.weekly_sccb_keys = set()

                    data = jira.search(self.jql_var.get(), self.max_results_var.get())

                issues = data.get("issues", [])
                total = data.get("total", len(issues))
                for it in issues:
                    f = it["fields"]
                    key = it["key"]
                    st = (f.get("status") or {}).get("name", "")

                    # SCCB 대상 검증값들은 Search 단계에서는 빈칸으로 두고,
                    # SCCB 대상 모드일 때만 백그라운드로 채운다.
                    body_len = "..."
                    rollout = "..."
                    links = "..."

                    self.selected[key] = False
                    self.issue_status[key] = st
                    iid = self.tree.insert("", END, values=(
                        "☐", key, f.get("summary", ""),
                        body_len,
                        rollout,
                        "...",  # err_table placeholder
                        links,
                        "...",  # tcgen placeholder
                        "...",  # aio_test placeholder
                        "...",  # pr_merge placeholder
                        st,
                        re.sub(r"\s*/.*$", "", (f.get("assignee") or {}).get("displayName", "")),  # / 이후 영문 제거
                        f.get("duedate", "") or ""
                    ))
                    self.iid_by_key[key] = iid

                # SCCB 대상 모드일 때만, 각 이슈의 추가 검증 컬럼들을 채운다.
                # VOC 완료처리는 목록만 조회하고 기존 검증 항목은 수행하지 않는다.
                if self.sccb_mode == "target" and issues:
                    import concurrent.futures
                    
                    # 진행 상황 추적
                    completed = [0]
                    total_count = len(issues)
                    
                    def update_progress():
                        self.set_status(f"{len(issues)} issues (검증 중: {completed[0]}/{total_count})")
                    
                    # 1단계: 모든 이슈의 core 데이터를 배치로 미리 조회
                    issue_keys_list = [it["key"] for it in issues]
                    core_cache = {}
                    try:
                        core_cache = jira.get_issues_core_batch(issue_keys_list)
                        self.log(f"Core 데이터 배치 조회 완료: {len(core_cache)}건")
                    except Exception as e:
                        self.log(f"Core 데이터 배치 조회 실패, 개별 조회로 진행: {e}")

                    def fetch_and_set(k: str, iid: str):
                        try:
                            # 1) 공통 데이터(Description/IssueLinks/IssueID) - 캐시 사용
                            core = core_cache.get(k)
                            if not core:
                                core = jira.get_issue_core(k)
                            
                            fields = (core.get('fields') or {})
                            desc = fields.get('description') or ''
                            issuelinks = fields.get('issuelinks') or []
                            issue_id = core.get('id') or ''

                            # 2) 네트워크 없이 계산 가능한 항목을 먼저 반영
                            rollout_status = None
                            err_status = None
                            missing = None
                            try:
                                rollout_status = jira.get_design_rollout_status(k, desc=desc) if hasattr(jira, 'get_design_rollout_status') else ('OK' if jira.get_design_rollout_ok(k, desc=desc) else 'FAIL')
                            except Exception:
                                rollout_status = None
                            try:
                                err_status = jira.get_error_table_status(k, desc=desc) if hasattr(jira, 'get_error_table_status') else ('OK' if jira.get_error_table_ok(k, desc=desc) else 'FAIL')
                            except Exception:
                                err_status = None
                            try:
                                missing = (jira.get_link_validation(k, links=issuelinks).get('missing') or [])
                            except Exception:
                                missing = None

                            def _apply_fast():
                                if iid not in self.iid_by_key.values():
                                    return
                                self.tree.set(iid, 'rollout', 'ERR' if rollout_status is None else rollout_status)
                                self.tree.set(iid, 'err_table', 'ERR' if err_status is None else err_status)
                                if missing is None:
                                    self.tree.set(iid, 'links', 'ERR')
                                elif missing:
                                    self.tree.set(iid, 'links', 'Missing: ' + ','.join(missing))
                                else:
                                    self.tree.set(iid, 'links', 'OK')
                                self._apply_zebra()

                            self.after(0, _apply_fast)

                            # 3) 느린 항목 - 원본 로직 그대로 유지
                            def _safe_call(fn, *args, **kwargs):
                                try:
                                    return fn(*args, **kwargs)
                                except Exception:
                                    return None

                            # 본문 길이 - 2회 재시도 (권한없음이면 즉시 중단)
                            body_len = None
                            for _ in range(2):
                                body_len = _safe_call(jira.get_body_length_string_from_ui, k)
                                if body_len:
                                    break
                                if body_len == 'N/A(권한없음)':
                                    break
                            
                            # TC, AIO, PR - 원본 그대로
                            tc_res = _safe_call(jira.get_tc_generation_check, k)
                            aio_res = _safe_call(jira.get_aio_test_validation, k)
                            pr_res = _safe_call(jira.get_pr_merge_status, k, issue_id)

                            def _apply_slow(body_len=body_len, tc_res=tc_res, aio_res=aio_res, pr_res=pr_res):
                                if iid not in self.iid_by_key.values():
                                    return
                                self.tree.set(iid, 'body_len', body_len if body_len is not None and body_len != '' else '확인불가')
                                tc_ok = bool((tc_res or {}).get('ok', False))
                                self.tree.set(iid, 'tcgen', 'OK' if tc_ok else 'FAIL')
                                aio_status = (aio_res or {}).get('status') or 'ERR'
                                self.tree.set(iid, 'aio_test', aio_status)
                                self.tree.set(iid, 'pr_merge', pr_res if pr_res is not None else 'ERR')
                                self._apply_zebra()

                            self.after(0, _apply_slow)

                        except Exception:
                            # 어떤 예외도 전체 진행을 멈추지 않게 방어
                            def _apply_err():
                                if iid not in self.iid_by_key.values():
                                    return
                                for col in ('body_len','rollout','err_table','links','tcgen','aio_test','pr_merge'):
                                    if not self.tree.set(iid, col):
                                        self.tree.set(iid, col, 'ERR')
                                self._apply_zebra()
                            self.after(0, _apply_err)
                        finally:
                            completed[0] += 1
                            self.after(0, update_progress)

                    # ThreadPoolExecutor로 병렬 처리 (10 workers)
                    def run_all_checks():
                        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                            futures = []
                            for k, iid in list(self.iid_by_key.items()):
                                future = executor.submit(fetch_and_set, k, iid)
                                futures.append(future)
                            concurrent.futures.wait(futures)

                    threading.Thread(target=run_all_checks, daemon=True).start()
                shown = len(issues)
                matched_cnt = 0
                if self.weekly_sccb_keys:
                    matched_cnt = sum(1 for it in issues if (it.get("key") or "") in self.weekly_sccb_keys)
                if isinstance(total, int) and total > shown:
                    base_msg = f"{shown} issues (표시: {shown}/{total})"
                else:
                    base_msg = f"{shown} issues"
                if self.weekly_sccb_keys:
                    base_msg += f" / 주간SCCB 일치 {matched_cnt}건"
                self.set_status(base_msg)
                self.after(0, self._apply_zebra)
            except Exception as e:
                self.log(f"Search 실패 - {e}")
                self.set_status("0 issues")

        threading.Thread(target=worker, daemon=True).start()


    def _calc_row_result(self, rollout: str, err_table: str, links: str, tcgen: str, aio_test: str, pr_merge: str) -> str:
        # "OK" 포함 여부로 판정
        if not rollout or "OK" not in (rollout or "").upper():
            return "공란"
        if not err_table or "OK" not in (err_table or "").upper():
            return "FAIL"
        if not links or "OK" not in (links or "").upper():
            return "FAIL"
        if not tcgen or "OK" not in (tcgen or "").upper():
            return "FAIL"
        if not aio_test or 'OK' not in (aio_test or '').upper():
            return 'FAIL'
        if not pr_merge or "MERGED" not in (pr_merge or "").upper():
            if pr_merge and "N/A" in pr_merge.upper():
                return "N/A"
            return "FAIL"
        return "OK"

    def on_export_excel(self):
        # 현재 Tree 결과를 Excel(xlsx)로 내보낸다.
        rows = []
        for iid in self.tree.get_children(""):
            v = self.tree.item(iid, "values")
            if not v or len(v) < 12:
                continue
            # v: sel,key,summary,body_len,rollout,err_table,links,tcgen,aio_test,pr_merge,status,assignee,duedate
            sel, key, summary, body_len, rollout, err_table, links, tcgen, aio_test, pr_merge, status, assignee, duedate = v[:13]
            result = self._calc_row_result(rollout, err_table, links, tcgen, aio_test, pr_merge)
            rows.append([result, key, summary, body_len, rollout, err_table, links, tcgen, aio_test, pr_merge, status, assignee, duedate])

        if not rows:
            Messagebox.show_info("내보낼 데이터가 없습니다.", title="안내")
            return

        ts = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
        default_name = f"sccb_validation_{ts}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=default_name,
            title="검증 결과 저장"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation"

        header = ["RESULT", "KEY", "SUMMARY", "본문 길이", "횡전개 표", "연관 에러", "이슈연결", "LLM TC 생성", "AIO Test", "P/R 병합", "STATUS", "ASSIGNEE", "기한일"]
        ws.append(header)

        # 헤더 스타일
        header_font = Font(bold=True)
        for col_idx, h in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font

        fail_fill = PatternFill("solid", fgColor="C00000")
        fail_font = Font(color="FFFFFF")

        for r in rows:
            ws.append(r)

        # FAIL 행 색칠
        #for row_idx in range(2, 2 + len(rows)):
        #    if (ws.cell(row=row_idx, column=1).value or "").strip().upper() == "FAIL":
        #        for col_idx in range(1, len(header) + 1):
        #           c = ws.cell(row=row_idx, column=col_idx)
        #          c.fill = fail_fill
        #           c.font = fail_font

        # AutoFilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)+1}"
        ws.freeze_panes = "A2"
    
        # Column width (대략)
        widths = [10, 16, 70, 12, 12, 12, 28, 12, 12, 10, 18, 18, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)
        self.log(f"Excel Export 완료: {path}")
        Messagebox.show_info("Export 완료", title="안내")

    def open_selected_issue(self):
        k = self._get_first_selected_key()
        if k:
            webbrowser.open(f"{self.base_url_var.get().rstrip('/')}/browse/{k}")
            return
        Messagebox.show_info("선택된 이슈가 없습니다.", title="안내")

    def on_process_selected_approval(self):
        keys = [k for k, v in self.selected.items() if v]
        if not keys:
            Messagebox.show_warning("처리할 이슈를 선택하세요.", title="선택 없음")
            return

        jira = self._make_client()
        wf = TransitionWorkflow(
            jira_client=jira,
            log_fn=self.log,
            sccb_mode_getter=lambda: self.sccb_mode
        )

        def worker():
            self.log(f"Approval 전이 시작: {keys}")
            for k in keys:
                try:
                    wf.process_issue_to_approval(k, cached_status=self.issue_status.get(k))
                    # 전이 성공 시 캐시/그리드 상태 갱신
                    try:
                        new_st = jira.get_issue_status(k)
                        self.issue_status[k] = new_st
                        iid = self.iid_by_key.get(k)
                        if iid:
                            self.after(0, lambda iid=iid, st=new_st: self.tree.set(iid, "status", st))
                    except Exception:
                        pass
                except Exception as e:
                    self.log(f"{k}: 실패 - {e}")
            self.log("Approval 전이 종료")

        threading.Thread(target=worker, daemon=True).start()

    def on_process_selected(self):
        keys = [k for k, v in self.selected.items() if v]
        if not keys:
            Messagebox.show_warning("처리할 이슈를 선택하세요.", title="선택 없음")
            return

        jira = self._make_client()
        wf = TransitionWorkflow(
            jira_client=jira,
            log_fn=self.log,
            sccb_mode_getter=lambda: self.sccb_mode
        )

        def worker():
            self.log(f"처리 시작: {keys}")
            for k in keys:
                try:
                    if self.sccb_mode == "voc_complete":
                        wf.process_voc_linked_issues_from_parent(k)
                    else:
                        wf.process_issue_to_complete(k, cached_status=self.issue_status.get(k))
                except Exception as e:
                    self.log(f"{k}: 실패 - {e}")
            self.log("처리 종료")

        threading.Thread(target=worker, daemon=True).start()